"""Export utilities for generating CSV, Excel, and PDF files"""

import csv
import io
import os
from datetime import datetime
from flask import Response, make_response


def get_export_logo_path(root_path, system):
    """
    Return filesystem path to the company logo for use in exports, or None if not available.
    root_path: Flask app root path (e.g. current_app.root_path)
    system: InventorySystem instance with get_all_system_settings()
    """
    try:
        if not system:
            return None
        settings = system.get_all_system_settings()
        logo_val = (settings.get('logo_path') or {}).get('value') or ''
        if not logo_val or logo_val.startswith('http'):
            return None
        # logo_val is like /static/logo.png
        fs_path = os.path.join(root_path, logo_val.lstrip('/'))
        return fs_path if os.path.isfile(fs_path) else None
    except Exception:
        return None

def export_to_csv(data, filename, headers=None):
    """
    Export data to CSV format
    
    Args:
        data: List of dictionaries or list of lists
        filename: Name for the downloaded file
        headers: Optional list of header names
    
    Returns:
        Flask Response object with CSV file
    """
    output = io.StringIO()
    
    if not data:
        # Empty dataset
        writer = csv.writer(output)
        if headers:
            writer.writerow(headers)
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    
    # Handle list of dictionaries
    if isinstance(data[0], dict):
        fieldnames = headers or list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(data)
    # Handle list of lists
    else:
        writer = csv.writer(output)
        if headers:
            writer.writerow(headers)
        writer.writerows(data)
    
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


def export_to_excel(data, filename, headers=None, sheet_name='Sheet1', logo_path=None):
    """
    Export data to Excel format (requires openpyxl or xlsxwriter).
    If logo_path (filesystem path) is provided, the logo is placed at the top left and data starts below it.

    Args:
        data: List of dictionaries or list of lists
        filename: Name for the downloaded file
        headers: Optional list of header names
        sheet_name: Name for the Excel sheet
        logo_path: Optional filesystem path to logo image for top-left of sheet

    Returns:
        Flask Response object with Excel file
    """
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Logo at top left: reserve rows and add image
        logo_rows = 0
        if logo_path and os.path.isfile(logo_path):
            try:
                from openpyxl.drawing.image import Image as XLImage
                img = XLImage(logo_path)
                img.width, img.height = 120, 60
                ws.add_image(img, 'A1')
                logo_rows = 5  # space for logo
            except Exception:
                pass

        # Add headers
        if isinstance(data[0], dict) if data else False:
            headers = headers or list(data[0].keys())

        start_row = 1 + logo_rows
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # Add data
        data_start = start_row + 1 if headers else start_row
        if data:
            if isinstance(data[0], dict):
                for row_idx, row_data in enumerate(data, data_start):
                    for col_idx, key in enumerate(headers or row_data.keys(), 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
            else:
                for row_idx, row_data in enumerate(data, data_start):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_to_csv(data, filename.replace('.xlsx', '.csv'), headers)


def prepare_asset_export_data(assets):
    """Prepare asset data for export"""
    export_data = []
    for asset_name, asset_data in assets.items():
        export_data.append({
            'Asset Name': asset_name,
            'Category': asset_data.get('category', ''),
            'Quantity': asset_data.get('quantity', 0),
            'Price': asset_data.get('price', 0),
            'Supplier': asset_data.get('supplier', ''),
            'Location': asset_data.get('location', ''),
            'Department': asset_data.get('department', ''),
            'Serial Number': asset_data.get('serial_number', ''),
            'Model': asset_data.get('model', ''),
            'Purchase Date': asset_data.get('purchase_date', ''),
            'Warranty Date': asset_data.get('warranty_date', ''),
            'Status': asset_data.get('status', 'Available'),
            'Responsible Officer': asset_data.get('responsible_officer', ''),
            'Province': asset_data.get('province_name', ''),
            'Island': asset_data.get('island', ''),
            'Unit/Section': asset_data.get('unit_section', ''),
            'Asset Category': asset_data.get('asset_category', ''),
            'LPO Number': asset_data.get('lpo_number', ''),
            'Asset Condition': asset_data.get('asset_condition', ''),
            'Asset Tag': asset_data.get('asset_tag', ''),
        })
    return export_data


def prepare_user_export_data(users):
    """Prepare user data for export"""
    export_data = []
    for user in users:
        export_data.append({
            'Username': user.get('username', ''),
            'Email': user.get('email', ''),
            'Groups': ', '.join(user.get('groups', [])),
            'Created': user.get('created_at', ''),
            'Last Login': user.get('last_login', ''),
        })
    return export_data


def prepare_maintenance_export_data(maintenance_records):
    """Prepare maintenance data for export"""
    export_data = []
    for record in maintenance_records:
        export_data.append({
            'Asset': record.get('asset_name', ''),
            'Maintenance Type': record.get('maintenance_type', ''),
            'Scheduled Date': record.get('scheduled_date', ''),
            'Completed Date': record.get('completed_date', ''),
            'Status': record.get('status', ''),
            'Cost': record.get('cost', 0),
            'Notes': record.get('notes', ''),
            'Performed By': record.get('performed_by', ''),
        })
    return export_data


def prepare_transaction_export_data(transactions):
    """Prepare transaction data for export"""
    export_data = []
    for transaction in transactions:
        export_data.append({
            'Date': transaction.get('date', ''),
            'Asset': transaction.get('asset_name', ''),
            'Type': transaction.get('transaction_type', ''),
            'User': transaction.get('username', ''),
            'From': transaction.get('from_user', ''),
            'To': transaction.get('to_user', ''),
            'Notes': transaction.get('notes', ''),
        })
    return export_data
