"""
Generate Excel template for asset import with sample data and instructions
"""
import sys
sys.path.insert(0, '/home/ubuntu/assetManagement/venv/lib/python3.13/site-packages')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()

# Create sheets
ws_instructions = wb.active
ws_instructions.title = "Instructions"
ws_template = wb.create_sheet("Template")
ws_sample = wb.create_sheet("Sample Data")

# ===== INSTRUCTIONS SHEET =====
ws_instructions.merge_cells('A1:D1')
ws_instructions['A1'] = 'Asset Import Template - Instructions'
ws_instructions['A1'].font = Font(size=18, bold=True, color="FFFFFF")
ws_instructions['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws_instructions['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_instructions.row_dimensions[1].height = 30

instructions = [
    ["", "", "", ""],
    ["How to Use This Template", "", "", ""],
    ["", "", "", ""],
    ["1. Choose Your Method:", "", "", ""],
    ["   • Use the 'Template' sheet for a blank import", "", "", ""],
    ["   • Use the 'Sample Data' sheet to see examples", "", "", ""],
    ["", "", "", ""],
    ["2. Fill in Your Data:", "", "", ""],
    ["   • Required fields: name, quantity, price", "", "", ""],
    ["   • Optional fields: All other columns", "", "", ""],
    ["   • Do NOT modify column headers", "", "", ""],
    ["", "", "", ""],
    ["3. Data Format Guidelines:", "", "", ""],
    ["   • name: Unique identifier (text)", "", "", ""],
    ["   • quantity: Number of items (integer)", "", "", ""],
    ["   • price: Cost per unit (decimal, e.g., 1250.00)", "", "", ""],
    ["   • purchase_date: Format YYYY-MM-DD (e.g., 2024-01-15)", "", "", ""],
    ["   • depreciation_method: 'straight_line' or 'declining_balance'", "", "", ""],
    ["   • useful_life_years: Number (e.g., 5)", "", "", ""],
    ["   • salvage_value: Decimal (e.g., 200.00)", "", "", ""],
    ["", "", "", ""],
    ["4. Import Process:", "", "", ""],
    ["   • Save your file as CSV or Excel (.xlsx)", "", "", ""],
    ["   • Go to Tools → Import Data in the system", "", "", ""],
    ["   • Select your file and click Import", "", "", ""],
    ["   • System will validate and import your data", "", "", ""],
    ["", "", "", ""],
    ["5. Tips:", "", "", ""],
    ["   • Remove sample data before importing your own", "", "", ""],
    ["   • Ensure supplier names exist in system", "", "", ""],
    ["   • Use consistent naming for departments/locations", "", "", ""],
    ["   • Keep backups of your import files", "", "", ""],
]

for idx, row in enumerate(instructions, start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws_instructions.cell(row=idx, column=col_idx, value=value)
        if idx == 2:
            cell.font = Font(size=14, bold=True, color="4472C4")

# Style instruction cells
for row in ws_instructions.iter_rows(min_row=4, max_row=32, min_col=1, max_col=4):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws_instructions.column_dimensions['A'].width = 50
ws_instructions.column_dimensions['B'].width = 20
ws_instructions.column_dimensions['C'].width = 20
ws_instructions.column_dimensions['D'].width = 20

# ===== FIELD DEFINITIONS =====
headers = [
    'name', 'quantity', 'price', 'description', 'low_stock_threshold',
    'category', 'supplier', 'department', 'location', 'model',
    'brand', 'serial_number', 'purchase_date', 'depreciation_method',
    'useful_life_years', 'salvage_value'
]

descriptions = [
    'Unique asset name (REQUIRED)',
    'Number of items (REQUIRED)',
    'Cost per unit (REQUIRED)',
    'Detailed description',
    'Alert threshold (default: 5)',
    'Asset category',
    'Supplier/vendor name',
    'Department assignment',
    'Physical location',
    'Model number',
    'Brand/manufacturer',
    'Serial number',
    'Purchase date (YYYY-MM-DD)',
    'straight_line or declining_balance',
    'Years of useful life',
    'Residual value after depreciation'
]

# ===== TEMPLATE SHEET =====
# Add headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, header in enumerate(headers, start=1):
    cell = ws_template.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws_template.column_dimensions[get_column_letter(col_idx)].width = 15

# Add descriptions row
desc_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
for col_idx, desc in enumerate(descriptions, start=1):
    cell = ws_template.cell(row=2, column=col_idx, value=desc)
    cell.fill = desc_fill
    cell.font = Font(italic=True, size=9)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = thin_border

ws_template.row_dimensions[2].height = 30

# Add empty rows for data entry
for row_idx in range(3, 23):
    for col_idx in range(1, len(headers) + 1):
        cell = ws_template.cell(row=row_idx, column=col_idx)
        cell.border = thin_border

# ===== SAMPLE DATA SHEET =====
# Add headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws_sample.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws_sample.column_dimensions[get_column_letter(col_idx)].width = 15

# Sample data
sample_data = [
    ['Dell Latitude 5520', 15, 1250.00, '15.6 inch business laptop with Intel i5 processor', 3, 'Electronics', 'Tech Suppliers Inc', 'IT Department', 'Building A - Floor 3', 'Latitude 5520', 'Dell', 'DL5520-2024-001', '2024-01-15', 'straight_line', 5, 200.00],
    ['HP LaserJet Pro M404dn', 5, 450.00, 'Monochrome laser printer with duplex printing', 2, 'Office Equipment', 'Office Depot', 'Administration', 'Building A - Floor 2', 'LaserJet Pro M404dn', 'HP', 'HP-LJ404-2024-001', '2024-02-20', 'straight_line', 7, 50.00],
    ['Herman Miller Aeron Chair', 25, 1200.00, 'Ergonomic office chair with lumbar support', 5, 'Furniture', 'Office Furniture Plus', 'Multiple Departments', 'Warehouse', 'Aeron', 'Herman Miller', 'HM-AERON-2024-001', '2024-01-10', 'straight_line', 10, 100.00],
    ['iPhone 14 Pro', 10, 999.00, 'Company smartphone with 256GB storage', 2, 'Mobile Devices', 'Apple Store', 'Sales Department', 'Building B - Floor 1', 'iPhone 14 Pro', 'Apple', 'APL-IP14P-2024-001', '2024-03-05', 'declining_balance', 3, 150.00],
    ['Cisco Catalyst 2960', 3, 2500.00, '24-port managed network switch', 1, 'Network Equipment', 'Network Solutions', 'IT Department', 'Server Room', 'Catalyst 2960', 'Cisco', 'CSC-CAT2960-2024-001', '2024-01-25', 'straight_line', 7, 300.00],
    ['Dell OptiPlex 7090', 20, 950.00, 'Desktop computer with Intel i7 and 16GB RAM', 4, 'Electronics', 'Tech Suppliers Inc', 'Accounting', 'Building A - Floor 4', 'OptiPlex 7090', 'Dell', 'DL-OPT7090-2024-001', '2024-02-01', 'straight_line', 5, 150.00],
    ['Samsung 27 4K Monitor', 30, 350.00, '27-inch 4K UHD monitor with USB-C', 5, 'Electronics', 'Tech Suppliers Inc', 'Multiple Departments', 'Warehouse', 'U28E590D', 'Samsung', 'SAM-U28E-2024-001', '2024-01-20', 'straight_line', 5, 50.00],
    ['Logitech MX Master 3', 50, 99.00, 'Wireless mouse for professionals', 10, 'Computer Accessories', 'Office Depot', 'Multiple Departments', 'Warehouse', 'MX Master 3', 'Logitech', 'LOG-MXM3-2024-001', '2024-02-15', 'straight_line', 3, 10.00],
    ['APC Smart-UPS 1500VA', 8, 650.00, 'Uninterruptible power supply with LCD display', 2, 'Power Equipment', 'APC Direct', 'IT Department', 'Server Room', 'SMT1500C', 'APC', 'APC-SMT1500-2024-001', '2024-01-30', 'straight_line', 8, 100.00],
    ['Microsoft Surface Pro 9', 12, 1100.00, '2-in-1 tablet with detachable keyboard', 3, 'Electronics', 'Microsoft Store', 'Executive Team', 'Building C - Floor 5', 'Surface Pro 9', 'Microsoft', 'MS-SP9-2024-001', '2024-03-10', 'straight_line', 4, 200.00],
]

# Add sample data rows
for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_sample.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [2, 5, 15]:  # Integer columns
            cell.number_format = '0'
        elif col_idx in [3, 16]:  # Decimal columns
            cell.number_format = '#,##0.00'
        cell.alignment = Alignment(vertical="top")

# Freeze top row in template and sample sheets
ws_template.freeze_panes = 'A2'
ws_sample.freeze_panes = 'A2'

# Save workbook
wb.save('/home/ubuntu/assetManagement/src/static/asset_import_template.xlsx')
print("Excel template created successfully: src/static/asset_import_template.xlsx")
